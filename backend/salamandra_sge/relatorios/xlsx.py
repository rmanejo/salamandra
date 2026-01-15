from io import BytesIO

from openpyxl import Workbook


def _as_bytes(workbook):
    buf = BytesIO()
    workbook.save(buf)
    return buf.getvalue()


def _safe_value(value):
    return "" if value is None else value


def pauta_turma_xlsx(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pauta"

    header = [
        "Numero",
        "Aluno",
        "T1_ACS1", "T1_ACS2", "T1_ACS3", "T1_MAP", "T1_MACS", "T1_ACP", "T1_MT", "T1_COM",
        "T2_ACS1", "T2_ACS2", "T2_ACS3", "T2_MAP", "T2_MACS", "T2_ACP", "T2_MT", "T2_COM",
        "T3_ACS1", "T3_ACS2", "T3_ACS3", "T3_MAP", "T3_MACS", "T3_ACP", "T3_MT", "T3_COM",
        "MFD",
    ]
    ws.append(header)

    for aluno in data["pauta"]:
        t1 = aluno["trimesters"][1]
        t2 = aluno["trimesters"][2]
        t3 = aluno["trimesters"][3]

        def _acs_list(tri):
            acs = tri.get("acs") or []
            return [
                _safe_value(acs[0] if len(acs) > 0 else None),
                _safe_value(acs[1] if len(acs) > 1 else None),
                _safe_value(acs[2] if len(acs) > 2 else None),
            ]

        row = [
            _safe_value(aluno.get("numero_turma")),
            aluno["nome"],
            *_acs_list(t1),
            _safe_value(t1.get("map")),
            _safe_value(t1.get("macs")),
            _safe_value(t1.get("acp")),
            _safe_value(t1.get("mt")),
            _safe_value(t1.get("com")),
            *_acs_list(t2),
            _safe_value(t2.get("map")),
            _safe_value(t2.get("macs")),
            _safe_value(t2.get("acp")),
            _safe_value(t2.get("mt")),
            _safe_value(t2.get("com")),
            *_acs_list(t3),
            _safe_value(t3.get("map")),
            _safe_value(t3.get("macs")),
            _safe_value(t3.get("acp")),
            _safe_value(t3.get("mt")),
            _safe_value(t3.get("com")),
            _safe_value(aluno.get("mfd")),
        ]
        ws.append(row)

    return _as_bytes(wb)


def pauta_turma_geral_xlsx(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pauta"

    disciplinas = data["disciplinas"]
    header = ["Numero", "Aluno"] + [d["nome"] for d in disciplinas] + ["Media_Final", "Situacao"]
    ws.append(header)

    for aluno in data["pauta"]:
        row = [
            _safe_value(aluno.get("numero_turma")),
            aluno["nome"],
        ]
        for disc in disciplinas:
            row.append(_safe_value(aluno["disciplinas"].get(disc["id"])))
        row.append(_safe_value(aluno.get("media_final")))
        row.append(aluno.get("situacao", ""))
        ws.append(row)

    return _as_bytes(wb)


def declaracao_aluno_xlsx(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Declaracao"

    header = ["Disciplina", "T1", "T2", "T3", "MFD", "Situacao"]
    ws.append(header)

    for disc in data["disciplinas"]:
        row = [
            disc["disciplina_nome"],
            _safe_value(disc["trimestres"].get(1)),
            _safe_value(disc["trimestres"].get(2)),
            _safe_value(disc["trimestres"].get(3)),
            _safe_value(disc.get("mfd")),
            disc.get("situacao", ""),
        ]
        ws.append(row)

    return _as_bytes(wb)


def situacao_academica_xlsx(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Situacao"

    header = [
        "Disciplina",
        "T1_ACS1", "T1_ACS2", "T1_ACS3", "T1_MAP", "T1_MACS", "T1_ACP", "T1_MT", "T1_COM",
        "T2_ACS1", "T2_ACS2", "T2_ACS3", "T2_MAP", "T2_MACS", "T2_ACP", "T2_MT", "T2_COM",
        "T3_ACS1", "T3_ACS2", "T3_ACS3", "T3_MAP", "T3_MACS", "T3_ACP", "T3_MT", "T3_COM",
        "MFD",
    ]
    ws.append(header)

    for disc in data["disciplinas"]:
        t1 = disc["trimesters"][1]
        t2 = disc["trimesters"][2]
        t3 = disc["trimesters"][3]

        def _acs_list(tri):
            acs = tri.get("acs") or []
            return [
                _safe_value(acs[0] if len(acs) > 0 else None),
                _safe_value(acs[1] if len(acs) > 1 else None),
                _safe_value(acs[2] if len(acs) > 2 else None),
            ]

        row = [
            disc["disciplina_nome"],
            *_acs_list(t1),
            _safe_value(t1.get("map")),
            _safe_value(t1.get("macs")),
            _safe_value(t1.get("acp")),
            _safe_value(t1.get("mt")),
            _safe_value(t1.get("com")),
            *_acs_list(t2),
            _safe_value(t2.get("map")),
            _safe_value(t2.get("macs")),
            _safe_value(t2.get("acp")),
            _safe_value(t2.get("mt")),
            _safe_value(t2.get("com")),
            *_acs_list(t3),
            _safe_value(t3.get("map")),
            _safe_value(t3.get("macs")),
            _safe_value(t3.get("acp")),
            _safe_value(t3.get("mt")),
            _safe_value(t3.get("com")),
            _safe_value(disc.get("mfd")),
        ]
        ws.append(row)

    return _as_bytes(wb)


def caderneta_xlsx(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Caderneta"

    header = [
        "Numero",
        "Aluno",
        "Sexo",
        "Status",
        "T1_ACS1", "T1_ACS2", "T1_ACS3", "T1_MAP", "T1_ACP", "T1_MACS", "T1_MT", "T1_COM",
        "T2_ACS1", "T2_ACS2", "T2_ACS3", "T2_MAP", "T2_ACP", "T2_MACS", "T2_MT", "T2_COM",
        "T3_ACS1", "T3_ACS2", "T3_ACS3", "T3_MAP", "T3_ACP", "T3_MACS", "T3_MT", "T3_COM",
        "MFD",
    ]
    ws.append(header)

    for row in data["rows"]:
        n1 = row["notas"]["1"]
        n2 = row["notas"]["2"]
        n3 = row["notas"]["3"]
        r1 = row["resumo"]["1"]
        r2 = row["resumo"]["2"]
        r3 = row["resumo"]["3"]

        ws.append([
            _safe_value(row.get("numero_turma")),
            row["nome_completo"],
            row.get("sexo", ""),
            row.get("status", ""),
            _safe_value(n1.get("ACS1")),
            _safe_value(n1.get("ACS2")),
            _safe_value(n1.get("ACS3")),
            _safe_value(n1.get("MAP")),
            _safe_value(n1.get("ACP")),
            _safe_value(r1.get("macs")),
            _safe_value(r1.get("mt")),
            _safe_value(r1.get("com")),
            _safe_value(n2.get("ACS1")),
            _safe_value(n2.get("ACS2")),
            _safe_value(n2.get("ACS3")),
            _safe_value(n2.get("MAP")),
            _safe_value(n2.get("ACP")),
            _safe_value(r2.get("macs")),
            _safe_value(r2.get("mt")),
            _safe_value(r2.get("com")),
            _safe_value(n3.get("ACS1")),
            _safe_value(n3.get("ACS2")),
            _safe_value(n3.get("ACS3")),
            _safe_value(n3.get("MAP")),
            _safe_value(n3.get("ACP")),
            _safe_value(r3.get("macs")),
            _safe_value(r3.get("mt")),
            _safe_value(r3.get("com")),
            _safe_value(row.get("mfd")),
        ])

    return _as_bytes(wb)


def lista_alunos_turma_xlsx(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Alunos"

    header = ["Numero", "Aluno", "Sexo", "Status"]
    ws.append(header)

    for aluno in data["alunos"]:
        ws.append([
            _safe_value(aluno.get("numero_turma")),
            aluno["nome"],
            aluno.get("sexo", ""),
            aluno.get("status", ""),
        ])

    return _as_bytes(wb)


def aprovados_reprovados_turma_xlsx(data):
    wb = Workbook()
    ws_aprovados = wb.active
    ws_aprovados.title = "Aprovados"
    ws_reprovados = wb.create_sheet("Reprovados")
    ws_sem = wb.create_sheet("Sem_Dados")

    header = ["Numero", "Aluno", "Media_Final", "Situacao"]
    for ws in (ws_aprovados, ws_reprovados, ws_sem):
        ws.append(header)

    def _append(ws, alunos):
        for aluno in alunos:
            ws.append([
                _safe_value(aluno.get("numero_turma")),
                aluno["nome"],
                _safe_value(aluno.get("media_final")),
                aluno.get("situacao", ""),
            ])

    _append(ws_aprovados, data["aprovados"])
    _append(ws_reprovados, data["reprovados"])
    _append(ws_sem, data["sem_dados"])

    return _as_bytes(wb)
