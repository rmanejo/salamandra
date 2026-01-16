import tempfile
from datetime import date
from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from core.models import CustomUser, District, School
from salamandra_sge.academico.models import Aluno, Classe, Disciplina, Professor, Turma, ProfessorTurmaDisciplina
from salamandra_sge.documentos.engine.caderneta import gerar_caderneta_documento
from salamandra_sge.documentos.models import DocumentTemplate, ProfessorProfile, TemplateMapping


def _make_template_file():
    wb = Workbook()
    ws = wb.active
    ws.title = "CADERNETA"
    ws["A1"] = "HEADER"
    ws["D2"] = "=SUM(A2:B2)"
    buf = BytesIO()
    wb.save(buf)
    return SimpleUploadedFile("template.xlsx", buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class CadernetaEngineTests(TestCase):
    def setUp(self):
        self.district = District.objects.create(name="Maputo")
        self.school = School.objects.create(name="Escola X", district=self.district)
        self.user = CustomUser.objects.create_user(
            email="prof@example.com",
            password="pass",
            first_name="Prof",
            last_name="Teste",
            role="PROFESSOR",
            school=self.school,
        )
        self.professor = Professor.objects.create(user=self.user, school=self.school)
        self.classe = Classe.objects.create(school=self.school, nome="10a")
        self.turma = Turma.objects.create(school=self.school, nome="A", classe=self.classe, ano_letivo=2024)
        self.disciplina = Disciplina.objects.create(school=self.school, nome="Matematica")
        ProfessorTurmaDisciplina.objects.create(
            school=self.school,
            professor=self.professor,
            turma=self.turma,
            disciplina=self.disciplina,
        )
        self.profile = ProfessorProfile.objects.create(
            professor=self.professor,
            area_formacao="Ciencias",
            nivel_academico="Licenciatura",
            contacto="1234",
        )

    def _add_alunos(self, total):
        alunos = []
        for i in range(total):
            alunos.append(
                Aluno.objects.create(
                    school=self.school,
                    nome_completo=f"Aluno {i}",
                    sexo="HOMEM" if i % 2 == 0 else "MULHER",
                    data_nascimento=date(2000, 1, 1),
                    turma_atual=self.turma,
                    classe_atual=self.classe,
                    numero_turma=i + 1,
                    ativo=True,
                )
            )
        return alunos

    def _create_template(self, faixa):
        template = DocumentTemplate.objects.create(
            school=self.school,
            doc_type="CADERNETA",
            faixa_alunos=faixa,
            template_file=_make_template_file(),
            version="1",
            is_active=True,
        )
        TemplateMapping.objects.create(
            template=template,
            sheet_name="CADERNETA",
            header_cells=TemplateMapping.default_header_mapping(),
            start_row_alunos=2,
            max_students=faixa,
            grade_columns={"ACS1": "A"},
            student_columns={"numero": "B", "nome": "C"},
            continuation_cell="A1",
        )
        return template

    def test_template_selection_by_range(self):
        self._create_template(50)
        template_75 = self._create_template(75)
        self._add_alunos(60)
        docs = gerar_caderneta_documento(
            user=self.user,
            turma_id=self.turma.id,
            disciplina_id=self.disciplina.id,
            trimestre=1,
            ano_lectivo=2024,
        )
        self.assertEqual(docs[0].template_used, template_75)

    def test_hide_unused_rows(self):
        self._create_template(50)
        self._add_alunos(1)
        docs = gerar_caderneta_documento(
            user=self.user,
            turma_id=self.turma.id,
            disciplina_id=self.disciplina.id,
            trimestre=1,
            ano_lectivo=2024,
        )
        wb = load_workbook(docs[0].file.path)
        ws = wb.active
        self.assertTrue(ws.row_dimensions[3].hidden)
        self.assertTrue(ws.row_dimensions[4].hidden)

    def test_continuation_parts(self):
        self._create_template(50)
        self._add_alunos(101)
        docs = gerar_caderneta_documento(
            user=self.user,
            turma_id=self.turma.id,
            disciplina_id=self.disciplina.id,
            trimestre=1,
            ano_lectivo=2024,
        )
        self.assertEqual(len(docs), 3)
        wb = load_workbook(docs[0].file.path)
        ws = wb.active
        self.assertEqual(ws["A1"].value, "Parte 1 de 3")

    def test_profile_incomplete_blocks_generation(self):
        self.profile.area_formacao = ""
        self.profile.save()
        client = APIClient()
        client.force_authenticate(self.user)
        response = client.post(
            "/api/documentos/caderneta/gerar/",
            {
                "turma_id": self.turma.id,
                "disciplina_id": self.disciplina.id,
                "trimestre": 1,
                "ano_lectivo": 2024,
            },
            format="json",
        )
        self.assertEqual(response.status_code, 403)

    def test_header_cells_filled(self):
        self._create_template(2)
        self._add_alunos(1)
        docs = gerar_caderneta_documento(
            user=self.user,
            turma_id=self.turma.id,
            disciplina_id=self.disciplina.id,
            trimestre=1,
            ano_lectivo=2024,
        )
        wb = load_workbook(docs[0].file.path)
        ws = wb.active
        self.assertEqual(ws["A9"].value, "Prof Teste")

    def test_formula_not_overwritten(self):
        self._create_template(2)
        self._add_alunos(1)
        docs = gerar_caderneta_documento(
            user=self.user,
            turma_id=self.turma.id,
            disciplina_id=self.disciplina.id,
            trimestre=1,
            ano_lectivo=2024,
        )
        wb = load_workbook(docs[0].file.path)
        ws = wb.active
        self.assertEqual(ws["D2"].value, "=SUM(A2:B2)")
