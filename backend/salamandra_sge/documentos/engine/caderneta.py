import logging
import math
import os
import re
from io import BytesIO
from types import SimpleNamespace

from django.core.files.base import ContentFile
from django.utils.text import slugify
from openpyxl import load_workbook

from salamandra_sge.academico.models import (
    Aluno,
    Disciplina,
    DirectorTurma,
    ProfessorTurmaDisciplina,
    Turma,
)
from salamandra_sge.avaliacoes.models import Nota
from salamandra_sge.relatorios.services import ReportService

from ..models import DocumentTemplate, GeneratedDocument, ProfessorProfile, TemplateMapping


logger = logging.getLogger(__name__)

TIPOS_NOTA = {"ACS1", "ACS2", "ACS3", "MAP", "ACP"}
DEFAULT_TEMPLATE_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    "relatorios",
    "templates",
)


def _resolve_cell(cell_or_col, row):
    if not cell_or_col:
        return None
    if re.match(r"^[A-Z]{1,3}[1-9][0-9]*$", cell_or_col):
        return cell_or_col
    return f"{cell_or_col}{row}"


def _safe_sheet(workbook, sheet_name):
    if sheet_name and sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    return workbook.worksheets[0]


def _clean_filename(value):
    slug = slugify(value or "")
    if not slug:
        return "documento"
    return slug.replace("-", "_")


def _default_template_path(faixa):
    candidates = {
        50: "caderneta_ate_50.xlsx",
        65: "caderneta_ate_65.xlsx",
        75: "caderneta_ate_75.xlsx",
        100: "caderneta_ate_100.xlsx",
    }
    filename = candidates.get(faixa)
    if not filename:
        return None
    path = os.path.join(DEFAULT_TEMPLATE_DIR, filename)
    return path if os.path.exists(path) else None


def _load_template_file(template, total_alunos):
    if template:
        return template.template_file.path, template, template.faixa_alunos
    faixas = [50, 65, 75, 100]
    available = [faixa for faixa in faixas if _default_template_path(faixa)]
    for faixa in available:
        if faixa >= total_alunos:
            path = _default_template_path(faixa)
            return path, None, faixa
    if available:
        faixa = max(available)
        path = _default_template_path(faixa)
        return path, None, faixa
    return None, None, None


def _select_template(school, doc_type, total_alunos):
    templates = (
        DocumentTemplate.objects.filter(
            school=school,
            doc_type=doc_type,
            is_active=True,
        )
        .order_by("faixa_alunos")
    )
    if not templates.exists():
        return None
    for template in templates:
        if template.faixa_alunos >= total_alunos:
            return template
    return templates.last()


def _build_header_values(
    profile,
    professor,
    turma,
    disciplina,
    ano_lectivo,
    total_alunos,
    alunos,
):
    homens = sum(1 for aluno in alunos if aluno.sexo == "HOMEM")
    mulheres = sum(1 for aluno in alunos if aluno.sexo == "MULHER")
    diretor = DirectorTurma.objects.filter(turma=turma).select_related("professor__user").first()

    return {
        "professor_nome": professor.user.get_full_name(),
        "habilitacoes_academicas": profile.nivel_academico,
        "area_formacao": profile.area_formacao,
        "contacto": profile.contacto,
        "disciplina": disciplina.nome,
        "classe": turma.classe.nome,
        "turma": turma.nome,
        "sala": getattr(turma, "sala", "") or "",
        "turno": getattr(turma, "turno", "") or "",
        "ano_lectivo": ano_lectivo,
        "total_alunos": total_alunos,
        "M": mulheres,
        "H": homens,
        "DT": diretor.professor.user.get_full_name() if diretor else "",
    }


def _collect_notes(school, turma, disciplina, ano_lectivo, trimestre):
    notas = Nota.objects.filter(
        school=school,
        turma=turma,
        disciplina=disciplina,
        ano_letivo=ano_lectivo,
        trimestre=trimestre,
    )
    notas_por_aluno = {}
    for nota in notas:
        notas_por_aluno.setdefault(nota.aluno_id, {})[nota.tipo] = nota.valor
    return notas_por_aluno


def _write_header(ws, mapping, values):
    for key, cell in (mapping.header_cells or {}).items():
        if cell and key in values:
            _safe_set_cell(ws, cell, values.get(key, ""))


def _safe_set_cell(ws, cell, value):
    if not cell:
        return
    for merged in ws.merged_cells.ranges:
        if cell in merged:
            cell = merged.start_cell.coordinate
            break
    ws[cell] = value


def _write_students(ws, mapping, alunos, notas_por_aluno, start_row, max_students):
    for idx, aluno in enumerate(alunos):
        row = start_row + idx
        for key, col in (mapping.student_columns or {}).items():
            cell = _resolve_cell(col, row)
            if not cell:
                continue
            value = ""
            if key == "numero":
                value = aluno.numero_turma
            elif key == "nome":
                value = aluno.nome_completo
            elif key == "sexo":
                value = aluno.sexo or ""
            elif key == "status":
                value = aluno.status or ""
            elif hasattr(aluno, key):
                value = getattr(aluno, key) or ""
            _safe_set_cell(ws, cell, value)

        notas = notas_por_aluno.get(aluno.id, {})
        for key, col in (mapping.grade_columns or {}).items():
            if key not in TIPOS_NOTA:
                continue
            cell = _resolve_cell(col, row)
            if not cell:
                continue
            _safe_set_cell(ws, cell, notas.get(key))

    used = len(alunos)
    for idx in range(used, max_students):
        row = start_row + idx
        ws.row_dimensions[row].hidden = True


def gerar_caderneta_documento(
    *,
    user,
    turma_id,
    disciplina_id,
    trimestre,
    ano_lectivo=None,
):
    turma = Turma.objects.select_related("classe").get(id=turma_id, school=user.school)
    disciplina = Disciplina.objects.get(id=disciplina_id, school=user.school)

    if not ReportService._can_view_caderneta(user, turma, disciplina):
        raise PermissionError("Sem permissão para gerar esta caderneta.")

    if user.role == "PROFESSOR":
        if not ProfessorTurmaDisciplina.objects.filter(turma=turma, disciplina=disciplina).exists():
            raise PermissionError("Disciplina não atribuída a esta turma.")

    if not hasattr(user, "docente_profile"):
        raise PermissionError("Perfil docente inválido.")

    profile, _ = ProfessorProfile.objects.get_or_create(professor=user.docente_profile)
    if not profile.is_complete:
        raise PermissionError("Perfil profissional incompleto. Complete o perfil antes de gerar documentos.")

    ano_lectivo = ano_lectivo or turma.ano_letivo
    alunos = list(
        Aluno.objects.filter(turma_atual=turma, ativo=True).order_by("numero_turma", "nome_completo")
    )
    total_alunos = len(alunos)

    template = _select_template(user.school, DocumentTemplate.DOC_TYPE_CADERNETA, total_alunos)
    template_path, template_record, faixa_fallback = _load_template_file(template, total_alunos)
    if not template_path:
        raise FileNotFoundError(
            f"Nenhum template ativo encontrado para CADERNETA na escola {user.school}."
        )

    mapping = template_record.mapping if template_record else None
    if not mapping:
        mapping = SimpleNamespace(
            header_cells=TemplateMapping.default_header_mapping(),
            sheet_name="",
            start_row_alunos=15,
            max_students=faixa_fallback or 50,
            grade_columns={},
            student_columns={
                "numero": "A",
                "nome": "B",
                "sexo": "C",
            },
            continuation_cell="",
        )

    notas_por_aluno = _collect_notes(user.school, turma, disciplina, ano_lectivo, trimestre)

    max_students = mapping.max_students
    parts_total = max(1, math.ceil(total_alunos / max_students))
    generated = []

    for part_index in range(parts_total):
        start = part_index * max_students
        end = start + max_students
        alunos_parte = alunos[start:end]

        workbook = load_workbook(template_path)
        ws = _safe_sheet(workbook, mapping.sheet_name)

        header_values = _build_header_values(
            profile=profile,
            professor=user.docente_profile,
            turma=turma,
            disciplina=disciplina,
            ano_lectivo=ano_lectivo,
            total_alunos=total_alunos,
            alunos=alunos,
        )
        _write_header(ws, mapping, header_values)

        if mapping.continuation_cell and parts_total > 1:
            ws[mapping.continuation_cell] = f"Parte {part_index + 1} de {parts_total}"

        _write_students(
            ws=ws,
            mapping=mapping,
            alunos=alunos_parte,
            notas_por_aluno=notas_por_aluno,
            start_row=mapping.start_row_alunos,
            max_students=max_students,
        )

        filename_base = (
            f"Caderneta_{_clean_filename(turma.nome)}_"
            f"{_clean_filename(disciplina.nome)}_{ano_lectivo}_T{trimestre}"
        )
        if parts_total > 1:
            filename_base += f"_P{part_index + 1}"
        filename = f"{filename_base}.xlsx"

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        doc = GeneratedDocument(
            school=user.school,
            doc_type=DocumentTemplate.DOC_TYPE_CADERNETA,
            turma=turma,
            disciplina=disciplina,
            trimestre=trimestre,
            ano_lectivo=ano_lectivo,
            generated_by=user,
            parts_total=parts_total,
            part_number=part_index + 1,
            template_used=template_record,
            template_version=template_record.version if template_record else "default",
        )
        doc.file.save(filename, ContentFile(output.read()), save=False)
        doc.save()
        generated.append(doc)

        logger.info(
            "Caderneta gerada: turma=%s disciplina=%s parte=%s/%s",
            turma.id,
            disciplina.id,
            part_index + 1,
            parts_total,
        )

    return generated
